from __future__ import annotations

import bisect
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# -----------------------------
# 환경설정
# -----------------------------
BASE_DIR = Path.cwd()
INPUT_PATTERNS = [
    "kb26q1.xlsx",
    "KB26q1.xlsx",
    "KB거래내역*.xlsx",
    "KB거래내역*.xlsm",
]
EXCHANGE_DIR_NAMES = ["exchange_rate", "Exchange_Rate", "EXCHANGE_RATE"]
OUTPUT_SUFFIX = "_정리.xlsx"

OUTPUT_COLUMNS = [
    "계좌번호",
    "계약자명",
    "PF명",
    "구분",
    "종목명",
    "매매일자",
    "수량",
    "매매단가",
    "매매금액",
    "위탁매매수수료",
    "각종세금",
]


# -----------------------------
# 환율 테이블
# -----------------------------
@dataclass
class ExchangeTable:
    dates_ord: List[int]
    rates: List[Decimal]

    def lookup(self, target_date: date) -> Decimal:
        ordinal = target_date.toordinal()
        pos = bisect.bisect_right(self.dates_ord, ordinal) - 1
        if pos < 0:
            raise ValueError(f"{target_date} 이전 환율이 없습니다.")
        return self.rates[pos]


# -----------------------------
# 공통 유틸
# -----------------------------
def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_date_safe(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        text = val.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def extract_account_info(a1_value: str) -> Tuple[str, str]:
    text = clean_text(a1_value)

    # 1) 하이픈 포함 계좌번호 우선
    m = re.search(r"(\d+(?:-\d+)+)", text)
    if m:
        account_no = m.group(1)
        holder = text[m.end():].strip()
        return account_no, holder

    # 2) fallback (숫자만 있는 경우)
    m2 = re.search(r"(\d{8,})", text)
    account_no = m2.group(1) if m2 else ""
    holder = re.sub(r"^[\d\-\s]+", "", text).strip()

    return account_no, holder


def find_exchange_dir(base_dir: Path) -> Optional[Path]:
    for name in EXCHANGE_DIR_NAMES:
        p = base_dir / name
        if p.exists() and p.is_dir():
            return p
    return None


def find_input_files(base_dir: Path) -> List[Path]:
    files: List[Path] = []
    for pattern in INPUT_PATTERNS:
        files.extend(base_dir.glob(pattern))
    return sorted({p.resolve() for p in files})


def adjust_fx_rate(currency_code: str, rate: Decimal) -> Decimal:
    code = clean_text(currency_code).upper()
    if code == "JPY":
        return rate / Decimal("100")
    return rate


def load_exchange_rates(exchange_dir: Optional[Path]) -> Dict[str, ExchangeTable]:
    rate_map: Dict[str, ExchangeTable] = {}
    if exchange_dir is None:
        return rate_map

    candidates: List[Path] = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(exchange_dir.glob(ext))
        candidates.extend(exchange_dir.glob(f"**/{ext}"))

    seen = set()
    unique_candidates = []
    for p in candidates:
        rp = str(p.resolve())
        if rp not in seen:
            seen.add(rp)
            unique_candidates.append(p)

    for file_path in unique_candidates:
        code = file_path.stem.strip().upper()
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        dates_ord: List[int] = []
        rates: List[Decimal] = []

        for row in range(10, ws.max_row + 1):
            d = parse_date_safe(ws[f"A{row}"].value)
            r = ws[f"C{row}"].value
            if d is None or r in (None, ""):
                continue
            dates_ord.append(d.toordinal())
            rates.append(to_decimal(r))

        wb.close()

        if not dates_ord:
            raise ValueError(f"환율 파일에서 데이터를 찾지 못했습니다: {file_path}")

        paired = sorted(zip(dates_ord, rates), key=lambda x: x[0])
        rate_map[code] = ExchangeTable(
            dates_ord=[x[0] for x in paired],
            rates=[x[1] for x in paired],
        )

    return rate_map


def get_effective_rate(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Decimal:
    """
    모든 환율은 반드시 exchange_rate 폴더의 환율 테이블에서 가져온다.
    원본 KB 파일의 '환율' 열은 참고하지 않는다.
    """
    currency = clean_text(row.get("통화구분")).upper()
    if currency in ("", "KRW"):
        return Decimal("1")

    trade_date = parse_date_safe(row.get("거래일자"))
    if trade_date is None:
        raise ValueError("거래일자를 해석할 수 없습니다.")

    table = fx_tables.get(currency)
    if table is None:
        available = ", ".join(sorted(fx_tables.keys())) if fx_tables else "없음"
        raise KeyError(
            f"통화코드 {currency} 에 해당하는 환율 파일을 찾지 못했습니다. 사용가능 코드: {available}"
        )

    fx = table.lookup(trade_date)
    return adjust_fx_rate(currency, fx)


def tax_sum(row: dict) -> Decimal:
    return (
        to_decimal(row.get("농특세/부가세"))
        + to_decimal(row.get("지방소득세"))
        + to_decimal(row.get("거래세 등"))
        + to_decimal(row.get("소득세"))
        + to_decimal(row.get("양도세"))
    )


def foreign_cash_delta(row: dict, prev_row: Optional[dict]) -> Decimal:
    current = to_decimal(row.get("외화예수금"))
    previous = to_decimal(prev_row.get("외화예수금")) if prev_row else Decimal("0")
    return current - previous


def calculate_row(
    row: dict, prev_row: Optional[dict], fx_tables: Dict[str, ExchangeTable]
) -> Tuple[Decimal, Decimal, Decimal, Decimal, Decimal, str]:
    tx = clean_text(row.get("거래종류"))
    qty = to_decimal(row.get("수량"))
    trade_amount = to_decimal(row.get("거래금액"))
    settle_amount = to_decimal(row.get("정산금액"))
    fx_settle_amount = to_decimal(row.get("외화정산금액"))
    unit_price = to_decimal(row.get("단가"))
    fee_domestic = to_decimal(row.get("수수료"))
    fee_foreign = to_decimal(row.get("국외수수료"))
    fx = get_effective_rate(row, fx_tables)
    delta_fx_cash = foreign_cash_delta(row, prev_row)

    # 1) 매수 / 매도 (해외주식)
    if tx in {"매수", "매도"}:
        out_qty = qty
        out_unit = unit_price * fx
        out_amount = out_qty * out_unit
        out_fee = fee_foreign * fx
        out_tax = to_decimal(row.get("거래세 등")) * fx
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 2) 주식장내매수 / 주식장내매도 (국내주식)
    if tx in {"주식장내매수", "주식장내매도"}:
        out_qty = qty
        out_unit = unit_price
        out_amount = out_qty * out_unit
        out_fee = fee_domestic
        out_tax = tax_sum(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 3) 외화매수 / 외화매도
    if tx in {"외화매수", "외화매도"}:
        out_qty = fx_settle_amount
        out_unit = fx
        out_amount = out_qty * out_unit
        out_fee = fee_domestic
        out_tax = (
            to_decimal(row.get("거래세 등"))
            + to_decimal(row.get("소득세"))
            + to_decimal(row.get("양도세"))
        )
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 4) 배당금 입금
    if tx == "배당금 입금":
        currency = clean_text(row.get("통화구분")).upper()
        out_fee = fee_domestic
        out_tax = (
            to_decimal(row.get("거래세 등"))
            + to_decimal(row.get("소득세"))
            + to_decimal(row.get("양도세"))
        )

        # 통화구분이 없는 경우
        if currency == "":
            out_qty = Decimal("0")
            out_unit = Decimal("0")
            out_amount = trade_amount
            return out_qty, out_unit, out_amount, out_fee, out_tax, ""

        # 통화구분이 있는 경우
        out_qty = fx_settle_amount
        out_unit = fx
        out_amount = out_qty * out_unit
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 5) 해외원천세 출금
    if tx == "해외원천세 출금":
        out_qty = fx_settle_amount
        out_unit = fx
        out_amount = Decimal("0")
        out_fee = Decimal("0")
        out_tax = out_qty * out_unit
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 6) 예탁금이용료 입금
    if tx == "예탁금이용료 입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount
        out_fee = Decimal("0")
        out_tax = (
            to_decimal(row.get("농특세/부가세"))
            + to_decimal(row.get("지방소득세"))
            + to_decimal(row.get("소득세"))
            + to_decimal(row.get("양도세"))
        )
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 7) 이자소득세추징 출금
    if tx == "이자소득세추징 출금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = Decimal("0")
        out_tax = (
            to_decimal(row.get("농특세/부가세"))
            + to_decimal(row.get("지방소득세"))
            + to_decimal(row.get("소득세"))
            + to_decimal(row.get("양도세"))
        )
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 8) 비용충당외화매수 출금
    if tx == "비용충당외화매수 출금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = settle_amount
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 9) ADR FEE 출금
    if tx == "ADR FEE 출금":
        prev_fx_cash = to_decimal(prev_row.get("외화예수금")) if prev_row else Decimal("0")
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = prev_fx_cash * fx
        out_tax = Decimal("0")
        note = ""
        if prev_row is None:
            note = "직전 거래가 없어 ADR FEE 계산 기준 외화예수금을 0으로 사용했습니다."
        return out_qty, out_unit, out_amount, out_fee, out_tax, note

    # 10) 대체입금
    if tx == "대체입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 11) 대체 입고 / 대체입고
    if tx in {"대체 입고", "대체입고"}:
        out_qty = qty
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 12) 외화계좌간대체 입금
    if tx == "외화계좌간대체 입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = delta_fx_cash * fx
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        note = ""
        if prev_row is None:
            note = "직전 거래가 없어 이전 외화예수금을 0으로 사용했습니다."
        return out_qty, out_unit, out_amount, out_fee, out_tax, note

    # 기본 fallback
    out_qty = qty
    out_unit = unit_price
    out_amount = trade_amount if trade_amount != 0 else settle_amount
    out_fee = fee_domestic
    out_tax = tax_sum(row)
    note = f"규칙 미지정 거래종류: {tx}"
    return out_qty, out_unit, out_amount, out_fee, out_tax, note


# -----------------------------
# 시트 처리
# -----------------------------
def read_sheet_rows(ws, fx_tables: Dict[str, ExchangeTable]) -> Tuple[List[List], List[str]]:
    account_no, holder = extract_account_info(ws["A1"].value)

    header_row_idx = 3
    headers = [ws.cell(header_row_idx, col).value for col in range(1, 40)]
    header_map = {clean_text(v): idx + 1 for idx, v in enumerate(headers) if v not in (None, "")}

    required_headers = [
        "거래일자",
        "거래종류",
        "수량",
        "거래금액",
        "정산금액",
        "외화정산금액",
        "거래세 등",
        "소득세",
        "양도세",
        "통화구분",
        "환율",
        "국외수수료",
        "종목명",
        "단가",
        "수수료",
        "농특세/부가세",
        "지방소득세",
        "외화예수금",
    ]
    for key in required_headers:
        if key not in header_map:
            raise KeyError(f"시트 '{ws.title}' 에서 필수 헤더 '{key}' 를 찾지 못했습니다.")

    output_rows: List[List] = []
    warnings: List[str] = []

    prev_row: Optional[dict] = None

    for r in range(4, ws.max_row + 1):
        trade_type = ws.cell(r, header_map["거래종류"]).value
        trade_date = ws.cell(r, header_map["거래일자"]).value

        if trade_type in (None, "") and trade_date in (None, ""):
            continue

        row = {key: ws.cell(r, col_idx).value for key, col_idx in header_map.items()}

        try:
            out_qty, out_unit, out_amount, out_fee, out_tax, note = calculate_row(row, prev_row, fx_tables)
        except Exception as exc:
            note = f"계산 실패({clean_text(row.get('거래종류'))}): {exc}"
            out_qty = Decimal("0")
            out_unit = Decimal("0")
            out_amount = Decimal("0")
            out_fee = Decimal("0")
            out_tax = Decimal("0")

        if note:
            warnings.append(f"[{ws.title} R{r}] {note}")

        output_rows.append([
            account_no,
            holder,
            "",
            clean_text(row.get("거래종류")),
            clean_text(row.get("종목명")),
            parse_date_safe(row.get("거래일자")),
            float(out_qty),
            float(out_unit),
            float(out_amount),
            float(out_fee),
            float(out_tax),
        ])

        prev_row = row

    return output_rows, warnings


# -----------------------------
# 저장
# -----------------------------
def autosize_columns(ws):
    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 28)


def save_output(output_path: Path, rows: List[List], warnings: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "정리"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[5].number_format = "yyyy-mm-dd"
        for idx in [6, 7, 8, 9, 10]:
            row[idx].number_format = "#,##0.00"

    autosize_columns(ws)

    if warnings:
        log_ws = wb.create_sheet("검토필요")
        log_ws.append(["메시지"])
        for msg in warnings:
            log_ws.append([msg])
        log_ws["A1"].font = Font(bold=True)
        log_ws.column_dimensions["A"].width = 120

    wb.save(output_path)


# -----------------------------
# 메인
# -----------------------------
def main():
    input_files = find_input_files(BASE_DIR)
    if not input_files:
        raise FileNotFoundError(
            f"작업폴더({BASE_DIR})에서 입력 파일을 찾지 못했습니다. "
            f"예상 파일명: kb26q1.xlsx"
        )

    exchange_dir = find_exchange_dir(BASE_DIR)
    if exchange_dir is None:
        raise FileNotFoundError(
            f"작업폴더({BASE_DIR}) 안에서 exchange_rate 폴더를 찾지 못했습니다. "
            f"허용 폴더명: {', '.join(EXCHANGE_DIR_NAMES)}"
        )

    fx_tables = load_exchange_rates(exchange_dir)

    for input_file in input_files:
        in_wb = load_workbook(input_file, data_only=True)
        all_rows: List[List] = []
        all_warnings: List[str] = []

        for sheet_name in in_wb.sheetnames:
            ws = in_wb[sheet_name]
            rows, warnings = read_sheet_rows(ws, fx_tables)
            all_rows.extend(rows)
            all_warnings.extend(warnings)

        output_path = input_file.with_name(f"{input_file.stem}{OUTPUT_SUFFIX}")
        save_output(output_path, all_rows, all_warnings)
        print(f"완료: {output_path}")
        if all_warnings:
            print(f"  - 검토필요 건수: {len(all_warnings)}")


if __name__ == "__main__":
    main()