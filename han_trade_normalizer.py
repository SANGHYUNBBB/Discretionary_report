
from __future__ import annotations

import bisect
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path.cwd()
INPUT_PATTERNS = [
    "han26q1.xlsx",
    "HAN26q1.xlsx",
    "한국투자증권*.xlsx",
    "한국투자증권*.xlsm",
]
EXCHANGE_DIR_NAMES = ["exchange_rate", "Exchange_Rate", "EXCHANGE_RATE"]
OUTPUT_SUFFIX = "_정리.xlsx"

OUTPUT_COLUMNS = [
    "계좌번호",
    "계약자명",
    "PF명",
    "구분",
    "종목명",
    "매매일자",
    "수량",
    "매매단가",
    "매매금액",
    "위탁매매수수료",
    "각종세금",
]

CURRENCY_CODES = {"USD", "JPY", "HKD", "CNY", "EUR", "GBP", "AUD", "CAD", "CHF"}


@dataclass
class ExchangeTable:
    dates_ord: List[int]
    rates: List[Decimal]

    def lookup(self, target_date: date) -> Decimal:
        ordinal = target_date.toordinal()
        pos = bisect.bisect_right(self.dates_ord, ordinal) - 1
        if pos < 0:
            raise ValueError(f"{target_date} 이전 환율이 없습니다.")
        return self.rates[pos]


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_date_safe(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = clean_text(val)
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def find_exchange_dir(base_dir: Path) -> Optional[Path]:
    for name in EXCHANGE_DIR_NAMES:
        p = base_dir / name
        if p.exists() and p.is_dir():
            return p
    return None


def find_input_files(base_dir: Path) -> List[Path]:
    files: List[Path] = []
    for pattern in INPUT_PATTERNS:
        files.extend(base_dir.glob(pattern))
    return sorted({p.resolve() for p in files})


def adjust_fx_rate(currency_code: str, rate: Decimal) -> Decimal:
    code = clean_text(currency_code).upper()
    if code == "JPY":
        return rate / Decimal("100")
    return rate


def load_exchange_rates(exchange_dir: Optional[Path]) -> Dict[str, ExchangeTable]:
    rate_map: Dict[str, ExchangeTable] = {}
    if exchange_dir is None:
        return rate_map

    candidates: List[Path] = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(exchange_dir.glob(ext))
        candidates.extend(exchange_dir.glob(f"**/{ext}"))

    seen = set()
    unique_candidates = []
    for p in candidates:
        rp = str(p.resolve())
        if rp not in seen:
            seen.add(rp)
            unique_candidates.append(p)

    for file_path in unique_candidates:
        code = file_path.stem.strip().upper()
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        dates_ord: List[int] = []
        rates: List[Decimal] = []

        for row in range(10, ws.max_row + 1):
            d = parse_date_safe(ws[f"A{row}"].value)
            r = ws[f"C{row}"].value
            if d is None or r in (None, ""):
                continue
            dates_ord.append(d.toordinal())
            rates.append(to_decimal(r))

        wb.close()

        if not dates_ord:
            raise ValueError(f"환율 파일에서 데이터를 찾지 못했습니다: {file_path}")

        paired = sorted(zip(dates_ord, rates), key=lambda x: x[0])
        rate_map[code] = ExchangeTable(
            dates_ord=[x[0] for x in paired],
            rates=[x[1] for x in paired],
        )

    return rate_map


def split_tx_and_currency(tx_raw: str) -> Tuple[str, str]:
    tx_raw = clean_text(tx_raw)
    suffix = tx_raw[-3:].upper() if len(tx_raw) >= 3 else ""
    if suffix in CURRENCY_CODES:
        return tx_raw[:-3].strip(), suffix
    return tx_raw, ""


def get_effective_rate(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Decimal:
    currency = clean_text(row.get("통화코드")).upper()
    if currency in ("", "KRW"):
        return Decimal("1")

    trade_date = parse_date_safe(row.get("거래일"))
    if trade_date is None:
        raise ValueError("거래일을 해석할 수 없습니다.")

    table = fx_tables.get(currency)
    if table is None:
        available = ", ".join(sorted(fx_tables.keys())) if fx_tables else "없음"
        raise KeyError(f"통화코드 {currency} 에 해당하는 환율 파일을 찾지 못했습니다. 사용가능 코드: {available}")

    fx = table.lookup(trade_date)
    return adjust_fx_rate(currency, fx)


def tax_sum_raw(row: dict) -> Decimal:
    return (
        to_decimal(row.get("거래세"))
        + to_decimal(row.get("세금"))
        + to_decimal(row.get("부가세"))
    )


def tax_sum_fx(row: dict, fx: Decimal) -> Decimal:
    return tax_sum_raw(row) * fx


def calculate_row(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Tuple[Decimal, Decimal, Decimal, Decimal, Decimal, str]:
    tx = clean_text(row.get("구분기준거래종류"))
    qty = to_decimal(row.get("거래수량"))
    trade_amount = to_decimal(row.get("거래금액"))
    unit_price = to_decimal(row.get("거래단가"))
    fee = to_decimal(row.get("수수료"))
    fx = get_effective_rate(row, fx_tables)

    if tx in {"해외증권매도", "해외증권매수"}:
        out_qty = qty
        out_unit = unit_price * fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "해외증권배당금입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount * fx
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "해외주식배당원화세금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "자동환전(외화매도)":
        out_qty = trade_amount
        out_unit = fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "외화예탁금이용료입금":
        out_qty = trade_amount
        out_unit = fx
        out_amount = out_qty * fx
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "외화예탁금이용료원화세금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "예탁금이용료":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "외화실시간직접매도환전":
        out_qty = trade_amount
        out_unit = fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "HTS자문사외화실시간직접매도환전":
        out_qty = trade_amount
        out_unit = fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "외화당사이체입금":
        out_qty = trade_amount
        out_unit = fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax_sum_fx(row, fx)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "랩대체계약출금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx == "당사이체출금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx.lower() == "smart+당사이체출금".lower():
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    if tx.lower() == "smart+당사이체입금".lower():
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = trade_amount
        out_fee = fee
        out_tax = tax_sum_raw(row)
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    out_qty = qty
    out_unit = unit_price
    out_amount = trade_amount
    out_fee = fee
    out_tax = tax_sum_raw(row)
    note = f"규칙 미지정 거래종류: {tx}"
    return out_qty, out_unit, out_amount, out_fee, out_tax, note


def build_header_map(ws) -> Dict[str, int]:
    headers = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    header_map = {}
    for idx, name in enumerate(headers, start=1):
        if name:
            if name not in header_map:
                header_map[name] = idx
            # 거래종류 텍스트가 실제로는 빈 헤더 다음 열(E열)에 있음
            if name == "거래종류" and idx < ws.max_column and clean_text(ws.cell(1, idx + 1).value) == "":
                header_map["거래종류_텍스트"] = idx + 1
    if "거래종류_텍스트" not in header_map:
        header_map["거래종류_텍스트"] = header_map["거래종류"]
    return header_map


def read_sheet_rows(ws, fx_tables: Dict[str, ExchangeTable]) -> Tuple[List[List], List[str]]:
    header_map = build_header_map(ws)

    required_headers = [
        "계좌번호", "계좌명", "거래일", "거래종류_텍스트",
        "종목명", "거래수량", "거래단가", "거래금액", "수수료",
        "거래세", "세금", "부가세"
    ]
    for key in required_headers:
        if key not in header_map:
            raise KeyError(f"시트 '{ws.title}' 에서 필수 헤더 '{key}' 를 찾지 못했습니다.")

    output_rows: List[List] = []
    warnings: List[str] = []

    current_account_no = ""
    current_holder = ""

    for r in range(2, ws.max_row + 1):
        tx_raw = clean_text(ws.cell(r, header_map["거래종류_텍스트"]).value)
        trade_date = ws.cell(r, header_map["거래일"]).value

        if tx_raw == "" and trade_date in (None, ""):
            continue

        account_no_val = clean_text(ws.cell(r, header_map["계좌번호"]).value)
        holder_val = clean_text(ws.cell(r, header_map["계좌명"]).value)
        if account_no_val:
            current_account_no = account_no_val
        if holder_val:
            current_holder = holder_val

        row = {}
        for key, col_idx in header_map.items():
            row[key] = ws.cell(r, col_idx).value

        base_tx, currency_code = split_tx_and_currency(tx_raw)
        row["통화코드"] = currency_code
        row["구분기준거래종류"] = base_tx

        try:
            out_qty, out_unit, out_amount, out_fee, out_tax, note = calculate_row(row, fx_tables)
        except Exception as exc:
            note = f"계산 실패({base_tx}): {exc}"
            out_qty = Decimal("0")
            out_unit = Decimal("0")
            out_amount = Decimal("0")
            out_fee = Decimal("0")
            out_tax = Decimal("0")

        if note:
            warnings.append(f"[{ws.title} R{r}] {note}")

        output_rows.append([
            current_account_no,
            current_holder,
            "",
            base_tx,
            clean_text(row.get("종목명")),
            parse_date_safe(row.get("거래일")),
            float(out_qty),
            float(out_unit),
            float(out_amount),
            float(out_fee),
            float(out_tax),
        ])

    return output_rows, warnings


def autosize_columns(ws):
    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 28)


def save_output(output_path: Path, rows: List[List], warnings: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "정리"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[5].number_format = "yyyy-mm-dd"
        for idx in [6, 7, 8, 9, 10]:
            row[idx].number_format = "#,##0.00"

    autosize_columns(ws)

    if warnings:
        log_ws = wb.create_sheet("검토필요")
        log_ws.append(["메시지"])
        for msg in warnings:
            log_ws.append([msg])
        log_ws["A1"].font = Font(bold=True)
        log_ws.column_dimensions["A"].width = 120

    wb.save(output_path)


def main():
    input_files = find_input_files(BASE_DIR)
    if not input_files:
        raise FileNotFoundError(
            f"작업폴더({BASE_DIR})에서 입력 파일을 찾지 못했습니다. 예상 파일명: han26q1.xlsx"
        )

    exchange_dir = find_exchange_dir(BASE_DIR)
    if exchange_dir is None:
        raise FileNotFoundError(
            f"작업폴더({BASE_DIR}) 안에서 exchange_rate 폴더를 찾지 못했습니다. "
            f"허용 폴더명: {', '.join(EXCHANGE_DIR_NAMES)}"
        )

    fx_tables = load_exchange_rates(exchange_dir)

    for input_file in input_files:
        in_wb = load_workbook(input_file, data_only=True)
        all_rows: List[List] = []
        all_warnings: List[str] = []

        for sheet_name in in_wb.sheetnames:
            ws = in_wb[sheet_name]
            rows, warnings = read_sheet_rows(ws, fx_tables)
            all_rows.extend(rows)
            all_warnings.extend(warnings)

        output_path = input_file.with_name(f"{input_file.stem}{OUTPUT_SUFFIX}")
        save_output(output_path, all_rows, all_warnings)
        print(f"완료: {output_path}")
        if all_warnings:
            print(f"  - 검토필요 건수: {len(all_warnings)}")


if __name__ == "__main__":
    main()
