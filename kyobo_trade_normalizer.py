
from __future__ import annotations

import bisect
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path.cwd()
INPUT_PATTERNS = [
    "kyobo26q1.xlsx",
    "KYOBO26q1.xlsx",
    "교보증권*.xlsx",
    "교보증권*.xlsm",
]
EXCHANGE_DIR_NAMES = ["exchange_rate", "Exchange_Rate", "EXCHANGE_RATE"]
OUTPUT_SUFFIX = "_정리.xlsx"

OUTPUT_COLUMNS = [
    "계좌번호",
    "계약자명",
    "PF명",
    "구분",
    "종목명",
    "매매일자",
    "수량",
    "매매단가",
    "매매금액",
    "위탁매매수수료",
    "각종세금",
]

FOREIGN_CODES = {"USD", "JPY", "HKD", "CNY", "EUR", "GBP", "AUD", "CAD", "CHF"}

SKIP_TYPES = {
    "해외주식매수입고",
    "해외주식매도출고",
    "타사대체입고신청",
}


@dataclass
class ExchangeTable:
    dates_ord: List[int]
    rates: List[Decimal]

    def lookup(self, target_date: date) -> Decimal:
        ordinal = target_date.toordinal()
        pos = bisect.bisect_right(self.dates_ord, ordinal) - 1
        if pos < 0:
            raise ValueError(f"{target_date} 이전 환율이 없습니다.")
        return self.rates[pos]


def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_date_safe(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = clean_text(val)
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def find_exchange_dir(base_dir: Path) -> Optional[Path]:
    for name in EXCHANGE_DIR_NAMES:
        p = base_dir / name
        if p.exists() and p.is_dir():
            return p
    return None


def find_input_files(base_dir: Path) -> List[Path]:
    files: List[Path] = []
    for pattern in INPUT_PATTERNS:
        files.extend(base_dir.glob(pattern))
    return sorted({p.resolve() for p in files})


def adjust_fx_rate(currency_code: str, rate: Decimal) -> Decimal:
    code = clean_text(currency_code).upper()
    if code == "JPY":
        return rate / Decimal("100")
    return rate


def load_exchange_rates(exchange_dir: Optional[Path]) -> Dict[str, ExchangeTable]:
    rate_map: Dict[str, ExchangeTable] = {}
    if exchange_dir is None:
        return rate_map

    candidates: List[Path] = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(exchange_dir.glob(ext))
        candidates.extend(exchange_dir.glob(f"**/{ext}"))

    seen = set()
    unique_candidates = []
    for p in candidates:
        rp = str(p.resolve())
        if rp not in seen:
            seen.add(rp)
            unique_candidates.append(p)

    for file_path in unique_candidates:
        code = file_path.stem.strip().upper()
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        dates_ord: List[int] = []
        rates: List[Decimal] = []

        for row in range(10, ws.max_row + 1):
            d = parse_date_safe(ws[f"A{row}"].value)
            r = ws[f"C{row}"].value
            if d is None or r in (None, ""):
                continue
            dates_ord.append(d.toordinal())
            rates.append(to_decimal(r))

        wb.close()

        if not dates_ord:
            raise ValueError(f"환율 파일에서 데이터를 찾지 못했습니다: {file_path}")

        paired = sorted(zip(dates_ord, rates), key=lambda x: x[0])
        rate_map[code] = ExchangeTable(
            dates_ord=[x[0] for x in paired],
            rates=[x[1] for x in paired],
        )

    return rate_map


def extract_account_info_from_a1(ws) -> Tuple[str, str]:
    text = clean_text(ws["A1"].value)

    # 계좌번호: 0000-00000-00 형태 우선 추출
    m_no = re.search(r"(\d{4}-\d{5}-\d{2})", text)
    account_no = m_no.group(1) if m_no else ""

    # 이름: 계좌번호 뒤의 나머지 문자열
    holder = ""
    if m_no:
        holder = text[m_no.end():].strip()
    else:
        # fallback
        m_name = re.search(r"([가-힣A-Za-z]+)\s*$", text)
        if m_name:
            holder = m_name.group(1)

    return account_no, holder


def find_header_row_and_map(ws) -> Tuple[int, Dict[str, int]]:
    for r in range(1, min(ws.max_row, 15) + 1):
        headers = [clean_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "적요명" in headers:
            return r, {headers[c - 1]: c for c in range(1, ws.max_column + 1) if headers[c - 1]}
    raise KeyError(f"시트 '{ws.title}' 에서 '적요명' 헤더를 찾지 못했습니다.")


def infer_currency(row: dict) -> str:
    for key in ["통화구분", "통화코드", "통화", "외화구분"]:
        val = clean_text(row.get(key)).upper()
        if val in FOREIGN_CODES:
            return val

    for key in ["적요명", "종목명"]:
        text = clean_text(row.get(key)).upper()
        for code in FOREIGN_CODES:
            if code in text:
                return code
    return ""


def get_effective_rate(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Decimal:
    currency = infer_currency(row)
    if currency in ("", "KRW"):
        return Decimal("1")

    trade_date = parse_date_safe(row.get("거래일자") or row.get("일자") or row.get("거래일"))
    if trade_date is None:
        raise ValueError("거래일자를 해석할 수 없습니다.")

    table = fx_tables.get(currency)
    if table is None:
        available = ", ".join(sorted(fx_tables.keys())) if fx_tables else "없음"
        raise KeyError(f"통화코드 {currency} 에 해당하는 환율 파일을 찾지 못했습니다. 사용가능 코드: {available}")

    fx = table.lookup(trade_date)
    return adjust_fx_rate(currency, fx)


def tax_value(row: dict) -> Decimal:
    for key in ["제세금", "세금", "거래세", "세금합계"]:
        if key in row:
            return to_decimal(row.get(key))
    return Decimal("0")


def fee_value(row: dict) -> Decimal:
    for key in ["수수료", "수수료금액"]:
        if key in row:
            return to_decimal(row.get(key))
    return Decimal("0")


def qty_value(row: dict) -> Decimal:
    for key in ["수량", "거래수량"]:
        if key in row:
            return to_decimal(row.get(key))
    return Decimal("0")


def unit_price_value(row: dict) -> Decimal:
    for key in ["단가", "거래단가"]:
        if key in row:
            return to_decimal(row.get(key))
    return Decimal("0")


def trade_amount_value(row: dict) -> Decimal:
    for key in ["거래금액", "금액"]:
        if key in row:
            return to_decimal(row.get(key))
    return Decimal("0")


def calculate_row(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Tuple[Optional[List], Optional[str]]:
    tx = clean_text(row.get("적요명"))
    if tx in SKIP_TYPES:
        return None, None

    qty = qty_value(row)
    unit = unit_price_value(row)
    amount = trade_amount_value(row)
    fee = fee_value(row)
    tax = tax_value(row)
    fx = get_effective_rate(row, fx_tables)
    currency = infer_currency(row)

    if tx in {"해외주식매도대금입금", "해외주식매수대금출금"}:
        out_qty = qty
        out_unit = unit * fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax * fx
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx == "배당금입금" and currency not in ("", "KRW"):
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount * fx
        out_fee = fee * fx
        out_tax = tax * fx
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx == "타사대체입고":
        out_qty = qty
        out_unit = unit * fx
        out_amount = out_qty * out_unit
        out_fee = fee * fx
        out_tax = tax * fx
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx == "정기예탁금이용료입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = fee
        out_tax = tax
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx in {"은행이체송금", "은행이체입금"}:
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = fee
        out_tax = tax
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    # 사용자가 지정한 원화 환전류 + 실제 파일에서 나온 외화 환전류 함께 반영
    if tx in {
        "원화입금(환전)",
        "원화출금(환전)",
        "원화출금(증거금환전)",
        "원화입금(증거금환전)",
        "외화출금(환전)",
        "외화입금(증거금환전)",
    }:
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = fee
        out_tax = tax
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx == "외화배당세금환급입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount * fx
        out_fee = fee * fx
        out_tax = tax * fx
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx == "배당소득세출금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = fee
        out_tax = tax
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    if tx == "권리정정출금(외화)":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount * fx
        out_fee = fee * fx
        out_tax = tax * fx
        return [out_qty, out_unit, out_amount, out_fee, out_tax], ""

    note = f"규칙 미지정 적요명: {tx}"
    out_qty = qty
    out_unit = unit
    out_amount = amount
    out_fee = fee
    out_tax = tax
    return [out_qty, out_unit, out_amount, out_fee, out_tax], note


def read_sheet_rows(ws, fx_tables: Dict[str, ExchangeTable]) -> Tuple[List[List], List[str]]:
    account_no, holder = extract_account_info_from_a1(ws)
    header_row_idx, header_map = find_header_row_and_map(ws)

    required = ["적요명", "거래일자"]
    for key in required:
        if key not in header_map:
            raise KeyError(f"시트 '{ws.title}' 에서 필수 헤더 '{key}' 를 찾지 못했습니다.")

    output_rows: List[List] = []
    warnings: List[str] = []

    for r in range(header_row_idx + 1, ws.max_row + 1):
        tx = clean_text(ws.cell(r, header_map["적요명"]).value)
        dt = ws.cell(r, header_map["거래일자"]).value

        if tx == "" and dt in (None, ""):
            continue

        row = {key: ws.cell(r, col).value for key, col in header_map.items()}

        try:
            calc, note = calculate_row(row, fx_tables)
        except Exception as exc:
            calc, note = [Decimal("0")] * 5, f"계산 실패({tx}): {exc}"

        if calc is None:
            continue

        if note:
            warnings.append(f"[{ws.title} R{r}] {note}")

        out_qty, out_unit, out_amount, out_fee, out_tax = calc

        output_rows.append([
            account_no,
            holder,
            "",
            tx,
            clean_text(row.get("종목명(거래상대명)")),
            parse_date_safe(row.get("거래일자")),
            float(out_qty),
            float(out_unit),
            float(out_amount),
            float(out_fee),
            float(out_tax),
        ])

    return output_rows, warnings


def autosize_columns(ws):
    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 28)


def save_output(output_path: Path, rows: List[List], warnings: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "정리"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[5].number_format = "yyyy-mm-dd"
        for idx in [6, 7, 8, 9, 10]:
            row[idx].number_format = "#,##0.00"

    autosize_columns(ws)

    if warnings:
        log_ws = wb.create_sheet("검토필요")
        log_ws.append(["메시지"])
        for msg in warnings:
            log_ws.append([msg])
        log_ws["A1"].font = Font(bold=True)
        log_ws.column_dimensions["A"].width = 120

    wb.save(output_path)


def main():
    input_files = find_input_files(BASE_DIR)
    if not input_files:
        raise FileNotFoundError(
            f"작업폴더({BASE_DIR})에서 입력 파일을 찾지 못했습니다. 예상 파일명: kyobo26q1.xlsx"
        )

    exchange_dir = find_exchange_dir(BASE_DIR)
    if exchange_dir is None:
        raise FileNotFoundError(
            f"작업폴더({BASE_DIR}) 안에서 exchange_rate 폴더를 찾지 못했습니다. "
            f"허용 폴더명: {', '.join(EXCHANGE_DIR_NAMES)}"
        )

    fx_tables = load_exchange_rates(exchange_dir)

    for input_file in input_files:
        in_wb = load_workbook(input_file, data_only=True)
        all_rows: List[List] = []
        all_warnings: List[str] = []

        for sheet_name in in_wb.sheetnames:
            ws = in_wb[sheet_name]
            rows, warnings = read_sheet_rows(ws, fx_tables)
            all_rows.extend(rows)
            all_warnings.extend(warnings)

        output_path = input_file.with_name(f"{input_file.stem}{OUTPUT_SUFFIX}")
        save_output(output_path, all_rows, all_warnings)
        print(f"완료: {output_path}")
        if all_warnings:
            print(f"  - 검토필요 건수: {len(all_warnings)}")


if __name__ == "__main__":
    main()
