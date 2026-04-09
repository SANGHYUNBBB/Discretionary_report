from __future__ import annotations

import bisect
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# -----------------------------
# 환경설정
# -----------------------------
BASE_DIR = Path.cwd()
INPUT_PATTERNS = [
    "Samsung26q1_*.xlsx",
    "Samsung26q1_*.xlsm",
    "삼성거래내역*.xlsx",
    "삼성거래내역*.xlsm",
]
EXCHANGE_DIR_NAMES = ["exchange_rate", "Exchange_Rate", "EXCHANGE_RATE"]
OUTPUT_SUFFIX = "_정리.xlsx"

OUTPUT_COLUMNS = [
    "계좌번호",
    "계약자명",
    "PF명",
    "구분",
    "종목명",
    "매매일자",
    "수량",
    "매매단가",
    "매매금액",
    "위탁매매수수료",
    "각종세금",
]

FOREIGN_STOCK_PATTERN = re.compile(
    r"^(미국\((?:NASDAQ|NYSE|AMEX)\)|홍콩|일본\(동경\)|상해\(후강퉁\)|심천\(선강퉁\))주식(매수|매도)$"
)
DOMESTIC_STOCK_TYPES = {"매수", "매도", "매수_NXT", "매도_NXT"}
CASH_TRANSFER_TYPES = {
    "이체입금",
    "대체입금",
    "대체출금",
    "이체출금",
    "출금",
    "랩대체입금",
}
SUPPORT_BONUS_TYPES = {"투자지원금", "투자지원금 입금"}
DIVIDEND_INBOUND_TYPES = {"배당입고", "청약입고", "상환입고", "대체입고"}
KNOWN_BUT_FALLBACK = {
    "청약",
    "청약입고",
    "상환입고",
    "대체입고",
    "외화이체입금",
    "시세이용료출금",
    "수수료입금",
    "자문사수수료입금",
    "자문사수수료출금",
}


@dataclass
class ExchangeTable:
    dates_ord: List[int]
    rates: List[Decimal]

    def lookup(self, target_date) -> Decimal:
        if target_date is None:
            raise ValueError("환율 조회를 위한 날짜가 없습니다.")
        ordinal = target_date.toordinal()
        pos = bisect.bisect_right(self.dates_ord, ordinal) - 1
        if pos < 0:
            raise ValueError(f"{target_date} 이전 환율이 없습니다.")
        return self.rates[pos]


# -----------------------------
# 공통 유틸
# -----------------------------
def to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(int(value))
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")

from datetime import datetime, date

def parse_date_safe(val):
    if val is None:
        return None

    if isinstance(val, date):
        return val

    if isinstance(val, datetime):
        return val.date()

    if isinstance(val, str):
        val = val.strip()

        # 여러 포맷 대응
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except:
                continue

    return None

def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_account_info(a1_value: str) -> Tuple[str, str]:
    text = clean_text(a1_value)

    # 1️⃣ 하이픈 포함 계좌번호 우선
    m_no = re.search(r"(\d+(?:-\d+)+)", text)
    if m_no:
        account_no = m_no.group(1)
    else:
        m2 = re.search(r"(\d{8,})", text)
        account_no = m2.group(1) if m2 else ""

    # 계약자명
    m_holder = re.search(r"\]\s*(.+)$", text)
    if m_holder:
        holder = m_holder.group(1).strip()
    elif m_no:
        holder = text[m_no.end():].strip()
    else:
        holder = ""

    return account_no, holder

def find_exchange_dir(base_dir: Path) -> Optional[Path]:
    for name in EXCHANGE_DIR_NAMES:
        p = base_dir / name
        if p.exists() and p.is_dir():
            return p
    return None


def find_input_files(base_dir: Path) -> List[Path]:
    files: List[Path] = []
    for pattern in INPUT_PATTERNS:
        files.extend(base_dir.glob(pattern))
    return sorted({p.resolve() for p in files})

def adjust_fx_rate(currency_code, rate):
    if not currency_code:
        return rate

    code = str(currency_code).strip().upper()

    if code == "JPY":
        return rate / Decimal("100")

    return rate
def load_exchange_rates(exchange_dir: Optional[Path]) -> Dict[str, ExchangeTable]:
    rate_map: Dict[str, ExchangeTable] = {}
    if exchange_dir is None:
        return rate_map

    candidates = []
    for ext in ("*.xlsx", "*.xlsm"):
        candidates.extend(exchange_dir.glob(ext))
        candidates.extend(exchange_dir.glob(f"**/{ext}"))

    seen = set()
    candidates = [p for p in candidates if not (str(p.resolve()) in seen or seen.add(str(p.resolve())))]

    for file_path in candidates:
        stem_upper = file_path.stem.strip().upper()
        m = re.search(r"([A-Z]{3})", stem_upper)
        code = m.group(1) if m else stem_upper
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        dates_ord: List[int] = []
        rates: List[Decimal] = []
        for row in range(10, ws.max_row + 1):
            date_value = ws[f"A{row}"].value
            rate_value = ws[f"C{row}"].value
            if date_value is None or rate_value in (None, ""):
                continue
            if hasattr(date_value, "date"):
                date_value = date_value.date()
            parsed_date = parse_date_safe(date_value)

            if parsed_date is not None:
                dates_ord.append(parsed_date.toordinal())
            rates.append(to_decimal(rate_value))

        if not dates_ord:
            raise ValueError(f"환율 파일에서 데이터를 찾지 못했습니다: {file_path}")

        paired = sorted(zip(dates_ord, rates), key=lambda x: x[0])
        rate_map[code] = ExchangeTable(
            dates_ord=[x[0] for x in paired],
            rates=[x[1] for x in paired],
        )
    return rate_map


# -----------------------------
# 계산 로직
# -----------------------------
def get_fx_rate(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Decimal:
    currency_code = clean_text(row.get("통화코드")).upper()
    trade_date = row.get("거래일자")

    if hasattr(trade_date, "date"):
        trade_date = trade_date.date()

    if currency_code in ("", "KRW"):
        return Decimal("1")

    table = fx_tables.get(currency_code)
    if table is None:
        available = ", ".join(sorted(fx_tables.keys())) if fx_tables else "없음"
        raise KeyError(f"통화코드 {currency_code} 에 해당하는 환율 파일을 찾지 못했습니다. 사용가능 코드: {available}")
    return table.lookup(trade_date)


def calculate_row(row: dict, fx_tables: Dict[str, ExchangeTable]) -> Tuple[Decimal, Decimal, Decimal, Decimal, Decimal, str]:
    tx = clean_text(row.get("거래명"))
    qty = to_decimal(row.get("거래수량"))
    unit = to_decimal(row.get("거래단가"))
    amount = to_decimal(row.get("거래금액"))
    fee = to_decimal(row.get("수수료/Fee"))
    tax_fee = to_decimal(row.get("제세금/대출이자"))
    foreign_fee = to_decimal(row.get("외화수수료"))

    # 1) 외화매수 / 외화매도
    if tx in {"외화매수", "외화매도"}:
        currency_code = clean_text(row.get("통화코드")).upper()
        out_qty = qty
        out_unit = adjust_fx_rate(currency_code, unit)
        out_amount = out_qty * out_unit
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 1-1) 외화이체입금 / 출금
    if tx in {"외화이체입금", "외화이체출금"}:
        currency_code = clean_text(row.get("통화코드")).upper()
        fx = get_fx_rate(row, fx_tables)
        fx = adjust_fx_rate(currency_code, fx)

        out_qty = qty
        out_unit = fx
        out_amount = out_qty * out_unit
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 2) 외국주식 매수/매도
    if FOREIGN_STOCK_PATTERN.match(tx):
        fx = get_fx_rate(row, fx_tables)
        currency_code = clean_text(row.get("통화코드")).upper()
        fx = adjust_fx_rate(currency_code, fx)

        out_qty = qty
        out_unit = unit * fx
        out_amount = out_qty * out_unit
        out_fee = foreign_fee * fx
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 3) 국내주식 매수/매도
    if tx in DOMESTIC_STOCK_TYPES:
        out_qty = qty
        out_unit = unit
        out_amount = out_qty * out_unit
        out_fee = fee
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 4) 세금출금(해외)
    if tx == "세금출금(해외)":
        currency_code = clean_text(row.get("통화코드")).upper()

        out_qty = qty
        out_unit = adjust_fx_rate(currency_code, unit)
        out_amount = Decimal("0")
        out_fee = Decimal("0")
        out_tax = out_qty * out_unit
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 5) 배당금입금
    if tx == "배당금입금":
        currency_code = clean_text(row.get("통화코드")).upper()
        out_fee = Decimal("0")
        out_tax = tax_fee

        if currency_code in ("", "KRW"):
            out_qty = Decimal("0")
            out_unit = Decimal("0")
            out_amount = amount
            return out_qty, out_unit, out_amount, out_fee, out_tax, ""

        fx = get_fx_rate(row, fx_tables)
        fx = adjust_fx_rate(currency_code, fx)

        out_qty = qty
        out_unit = fx
        out_amount = out_qty * out_unit
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 6) 이체입금 / 대체입금 / 대체출금 / 출금류
    if tx in CASH_TRANSFER_TYPES:
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 7) 투자지원금
    if tx in SUPPORT_BONUS_TYPES:
        out_qty = qty
        out_unit = unit
        out_amount = out_qty * out_unit
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 8) 이용료입금
    if tx == "이용료입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = Decimal("0")
        out_tax = fee + tax_fee
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 8-1) 수수료입금
    if tx == "수수료입금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = Decimal("0")
        out_fee = amount
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 8-2) 자문사수수료출금
    if tx == "자문사수수료출금":
        out_qty = Decimal("0")
        out_unit = Decimal("0")
        out_amount = amount
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    # 9) 배당입고
    if tx in DIVIDEND_INBOUND_TYPES:
        fx = get_fx_rate(row, fx_tables)
        currency_code = clean_text(row.get("통화코드")).upper()
        fx = adjust_fx_rate(currency_code, fx)

        out_qty = qty
        out_unit = unit * fx
        out_amount = out_qty * out_unit
        out_fee = Decimal("0")
        out_tax = Decimal("0")
        return out_qty, out_unit, out_amount, out_fee, out_tax, ""

    out_qty = qty
    out_unit = unit
    out_amount = amount
    out_fee = fee
    out_tax = tax_fee
    note = f"규칙 미지정 거래명: {tx}"
    return out_qty, out_unit, out_amount, out_fee, out_tax, note


# -----------------------------
# 시트 처리
# -----------------------------
def read_sheet_rows(ws, fx_tables: Dict[str, ExchangeTable]) -> Tuple[List[List], List[str]]:
    account_no, holder = extract_account_info(ws["A1"].value)

    header_row_idx = 3
    headers = [ws.cell(header_row_idx, col).value for col in range(1, ws.max_column + 1)]
    header_map = {clean_text(v): idx + 1 for idx, v in enumerate(headers) if v is not None}

    required_headers = [
        "거래일자",
        "거래명",
        "종목명",
        "거래수량",
        "거래단가",
        "거래금액",
        "제세금/대출이자",
        "수수료/Fee",
        "통화코드",
        "외화수수료",
    ]
    for key in required_headers:
        if key not in header_map:
            raise KeyError(f"시트 '{ws.title}' 에서 필수 헤더 '{key}' 를 찾지 못했습니다.")

    output_rows: List[List] = []
    warnings: List[str] = []

    for r in range(4, ws.max_row + 1):
        trade_name = ws.cell(r, header_map["거래명"]).value
        trade_date = ws.cell(r, header_map["거래일자"]).value
        if trade_name in (None, "") and trade_date in (None, ""):
            continue

        row = {key: ws.cell(r, col_idx).value for key, col_idx in header_map.items()}

        try:
            out_qty, out_unit, out_amount, out_fee, out_tax, note = calculate_row(row, fx_tables)
        except Exception as exc:
            note = f"계산 실패({clean_text(row.get('거래명'))}): {exc}"
            out_qty = to_decimal(row.get("거래수량"))
            out_unit = to_decimal(row.get("거래단가"))
            out_amount = to_decimal(row.get("거래금액"))
            out_fee = to_decimal(row.get("수수료/Fee"))
            out_tax = to_decimal(row.get("제세금/대출이자"))

        if note:
            warnings.append(f"[{ws.title} R{r}] {note}")

        output_rows.append([
            account_no,
            holder,
            "",  # PF명 공란
            clean_text(row.get("거래명")),
            clean_text(row.get("종목명")),
            row.get("거래일자"),
            float(out_qty),
            float(out_unit),
            float(out_amount),
            float(out_fee),
            float(out_tax),
        ])

    return output_rows, warnings


# -----------------------------
# 저장
# -----------------------------
def autosize_columns(ws):
    for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 28)


def save_output(output_path: Path, rows: List[List], warnings: List[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "정리"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 날짜/숫자 포맷
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[5].number_format = "yyyy-mm-dd"  # 매매일자
        for idx in [6, 7, 8, 9, 10]:
            row[idx].number_format = "#,##0.00"

    autosize_columns(ws)

    if warnings:
        log_ws = wb.create_sheet("검토필요")
        log_ws.append(["메시지"])
        for msg in warnings:
            log_ws.append([msg])
        log_ws["A1"].font = Font(bold=True)
        log_ws.column_dimensions["A"].width = 120

    wb.save(output_path)


# -----------------------------
# 메인
# -----------------------------
def main():
    base_dir = BASE_DIR
    input_files = find_input_files(base_dir)
    if not input_files:
        raise FileNotFoundError(
            f"작업폴더({base_dir})에서 입력 파일을 찾지 못했습니다. "
            f"예상 패턴: {', '.join(INPUT_PATTERNS)}"
        )

    exchange_dir = find_exchange_dir(base_dir)
    if exchange_dir is None:
        raise FileNotFoundError(
            f"작업폴더({base_dir}) 안에서 exchange_rate 폴더를 찾지 못했습니다. "
            f"허용 폴더명: {', '.join(EXCHANGE_DIR_NAMES)}"
        )
    fx_tables = load_exchange_rates(exchange_dir)

    for input_file in input_files:
        in_wb = load_workbook(input_file, data_only=True)
        all_rows: List[List] = []
        all_warnings: List[str] = []

        for sheet_name in in_wb.sheetnames:
            ws = in_wb[sheet_name]
            rows, warnings = read_sheet_rows(ws, fx_tables)
            all_rows.extend(rows)
            all_warnings.extend(warnings)

        output_path = input_file.with_name(f"{input_file.stem}{OUTPUT_SUFFIX}")
        save_output(output_path, all_rows, all_warnings)
        print(f"완료: {output_path}")
        if all_warnings:
            print(f"  - 검토필요 건수: {len(all_warnings)}")


if __name__ == "__main__":
    main()
